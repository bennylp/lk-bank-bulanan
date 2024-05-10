#!/usr/bin/env python3
from bs4 import BeautifulSoup
import glob
import json
import numpy as np
import openpyxl as oxl
import os
import pandas as pd
import re
import requests
import sys
from typing import List


def parse_text_file(spec: pd.DataFrame, input_path: str) -> pd.Series:
    print(f'Processing {input_path}')

    with open(input_path) as f:
        lines = f.readlines()

    fname = os.path.basename(input_path).split('.')[0]
    fts = pd.Timestamp(fname)

    last_lineno = 0

    data = { 
        'date': f'{fts.strftime("%Y-%m-%d")}',
    }
    for idx, row in spec.iterrows():
        pattern = row['pattern']
        target = row['target']
        force_sign = float(row['force_sign'])
        token_pos = row['quarterly_token_pos'] if fts.month in [3,6,9,12] else row['monthly_token_pos']
        mandatory = row['quarterly_mandatory'] if fts.month in [3,6,9,12] else row['monthly_mandatory']
        test_date0 = row['test_date0']
        test_value0 = row['test_value0']
        test_date1 = row['test_date1']
        test_value1 = row['test_value1']

        if not pattern:
            continue

        if not pd.isnull(test_date0) and test_date0:
            test_date0 = pd.Timestamp(test_date0)
        else:
            test_date0 = None
        if not pd.isnull(test_value0) and str(test_value0) != '':
            test_value0 = float(test_value0)
        else:
            test_value0 = np.NaN

        if not pd.isnull(test_date1) and test_date1:
            test_date1 = pd.Timestamp(test_date1)
        else:
            test_date1 = None
        if not pd.isnull(test_value1) and str(test_value1) != '':
            test_value1 = float(test_value1)
        else:
            test_value1 = np.NaN

        if pattern=='assign_value':
            try:
                data[target] = float(row['test_value0'])
            except:
                data[target] = row['test_value0']
            continue

        #if pattern=='Berdasarkan Kolektibilitas:':
        #    pattern = pattern

        lineno = last_lineno
        while lineno < len(lines):
            line = lines[lineno][:-1]  # strip ending '\n'
            if re.search(pattern, line) is not None:
                # Found the line
                if not pd.isnull(target) and target:
                    tokens = line.split()
                    value = tokens[token_pos]
                    if '(' in value:
                        sign = -1
                        value = value.replace('(', '').replace(')', '')
                    else:
                        sign = 1
                    if ',' in value:
                        raise RuntimeError(f'Check "{pattern}" value in {input_path}. It contains comma')
                    
                    if value=='-':
                        value = np.NaN
                    else:
                        value = float(value.replace('.', '').replace(',', '.')) * sign

                    if force_sign < 0 and value > 0:
                        value = -value
                    elif force_sign > 0 and value < 0:
                        value = -value

                    if target not in data or pd.isnull(data[target]):
                        data[target] = value
                    elif pd.isnull(value):
                        pass # No change
                    else:
                        data[target] = data[target] + value

                    if not pd.isnull(test_value0) and fts==test_date0:
                        if pd.isnull(test_value0) and not pd.isnull(data[target]) and data[target]:
                            raise ValueError(f'Field "{target}" mismatch: expected NaN or zero, got: {data[target]}, pattern: "{pattern}"')
                        if data[target] != test_value0:
                            raise ValueError(f'Field "{target}" mismatch: expected: {test_value0}, got: {data[target]}, pattern: "{pattern}"')

                    if not pd.isnull(test_value1) and fts==test_date1:
                        if data[target] != test_value1:
                            raise ValueError(f'Field "{target}" mismatch: expected: {test_value1}, got: {data[target]}, pattern: "{pattern}"')

                    # Do not add line number, same pattern may reuse the line with value on
                    # different position
                else:
                    lineno += 1

                break
            else:
                lineno += 1

        if lineno >= len(lines):
            err_msg = f'{input_path}: Unable to find pattern "{pattern}"'
            if mandatory:
                raise RuntimeError(err_msg)
            else:
                #print(f'Warning: {err_msg}')
                if not pd.isnull(target) and target:
                    data[target] = np.NaN
            # Don't update last_lineno to let search continue from the last found pattern
        else:
            last_lineno = lineno

    row = pd.Series(data)
    return row


def get_cell_value(cell: oxl.cell.cell.Cell):
    if '#' in cell.number_format:
        negative_format = -1 if '(' in cell.number_format[:2] else 1
        value = cell.value or 0.0
        return value * negative_format
    elif cell.number_format=='General':
        return cell.value or ''  # cell.value can be None
    else:
        raise RuntimeError(f'Unsupported cell number format "{cell.number_format}"')
    

def get_group_value(ws: oxl.worksheet.worksheet.Worksheet,
                    start_name_cell: oxl.cell.cell.Cell,
                    start_value_cell: oxl.cell.cell.Cell) -> float:
    values = [  get_cell_value(start_value_cell) or 0.0 ]
    name_cell = ws.cell( start_name_cell.row+1, start_name_cell.col_idx )
    value_cell = ws.cell( start_value_cell.row+1, start_value_cell.col_idx )
    while name_cell.alignment.indent > start_name_cell.alignment.indent:
        values.append( get_cell_value(value_cell) or 0.0 )
        name_cell = ws.cell( name_cell.row+1, name_cell.col_idx )
        value_cell = ws.cell( value_cell.row+1, value_cell.col_idx )
    return sum(values)


def parse_xlsx_file(spec: pd.DataFrame, input_path: str) -> pd.Series:
    wb = oxl.load_workbook(input_path)
    datas = {}

    for _, row in spec.iterrows():
        sheet = str(row['sheet'])
        pattern = row['pattern']
        src_operation = row['src_operation']
        target = row['target']
        force_sign = float(row['force_sign'])
        min_indent = row['min_indent']
        test_date0 = row['test_date0']
        test_value0 = row['test_value0']

        if not pattern:
            continue

        if pd.isnull(row['pattern_col']) or not row['pattern_col']:
            pattern_col = 1
        else:
            pattern_col = ord( row['pattern_col'] ) - 64  # one based
            assert pattern_col > 0 and pattern_col < 10

        if pd.isnull(row['value_col']) or not row['value_col']:
            value_col = pattern_col + 1
        else:
            value_col = ord( row['value_col'] ) - 64  # one based

        if not pd.isnull(test_date0) and test_date0:
            test_date0 = pd.Timestamp(test_date0)
        else:
            test_date0 = None
        if not pd.isnull(test_value0) and str(test_value0) != '':
            test_value0 = float(test_value0)
        else:
            test_value0 = np.NaN

        if pd.isnull(src_operation) or not src_operation:
            src_operation = 'get'

        if pattern=='assign_value':
            try:
                datas[target] = float(row['test_value0'])
            except:
                datas[target] = row['test_value0']

            if target=='pendapatan bersih':
                datas[target] = datas['pendapatan bunga'] - datas['beban bunga']

            continue

        if sheet not in wb.sheetnames:
            datas[target] = np.NaN
            continue

        ws = wb[ sheet ]

        pattern_cell = None
        if 'Beban operasional lain' in pattern:
            pass
        for cells in ws.iter_rows(min_col=pattern_col, max_col=pattern_col):
            cell = cells[0]
            if re.search(pattern, str(cell.value)) and (pd.isnull(min_indent) or cell.alignment.indent >= min_indent):
                pattern_cell = cell
                break
        
        if pattern_cell is None:
            raise RuntimeError(f'Unable to find pattern "{pattern}" in sheet {row["sheet"]} col {chr(64+pattern_col)}')

        start_cell = ws.cell( pattern_cell.row, value_col )
        if src_operation=='get':
            value = get_cell_value(start_cell)
        elif src_operation=='group':
            value = get_group_value(ws, pattern_cell, start_cell)
        else:
            raise ValueError(f'Invalid src_operation "{src_operation}" in spec for target "{target}"')
        
        if force_sign < 0 and value > 0:
            value = -value
        elif force_sign > 0 and value < 0:
            value = -value

        if target not in datas or pd.isnull(datas[target]):
            datas[target] = value
        elif pd.isnull(value):
            pass # No change
        else:
            datas[target] = datas[target] + value

        if not pd.isnull(test_value0) and pd.Timestamp(datas['date'])==pd.Timestamp(test_date0):
            if datas[target] != test_value0:
                raise ValueError(f'Field "{target}" mismatch: expected: {test_value0}, got: {datas[target]}, pattern: "{pattern}"')


    row = pd.Series(datas)
    assert row.index[0]=='date'
    assert row.index[1]=='pembulatan'

    # Ada laporan keuangan yang ditulis dalam Rp
    if row.iloc[2] > 1e9:
        row.iloc[2:] /= 1e6

    return row



def update_text_files(ticker: str, lktype: str):
    """
    lktype: None/empty/"konsol"/"induk"/whaterver
    """
    files = glob.glob(f'{ticker}/20*.txt')
    files = sorted(files)

    if lktype:
        spec = pd.read_excel(f'{ticker}/spec-{lktype}.xlsx')
        output_path = f'{ticker}/{ticker}-{lktype}.xlsx'
    else:
        spec = pd.read_excel(f'{ticker}/spec.xlsx')
        output_path = f'{ticker}/{ticker}.xlsx'

    spec['force_sign'] = spec['force_sign'].fillna(0).astype(int)
    spec['monthly_token_pos'] = spec['monthly_token_pos'].fillna(0).astype(int)
    spec['quarterly_token_pos'] = spec['quarterly_token_pos'].fillna(0).astype(int)
    spec['monthly_mandatory'] = spec['monthly_mandatory'].fillna(1).astype(int)
    spec['quarterly_mandatory'] = spec['quarterly_mandatory'].fillna(1).astype(int)
    
    if os.path.exists(output_path):
        output = pd.read_excel(output_path, parse_dates=[]).set_index('date', drop=True)
    else:
        output = None

    new_rows = []
    for file in files:
        date = os.path.basename(file).split('.')[0]
        if output is not None and date in output.index:
            continue

        row = parse_text_file(spec, file)
        if output is not None and row['date'] in output.index:
            # Already exists
            pass
        else:
            print(f'{output_path}: adding LK {row["date"]}')
            new_rows.append(row)

    if new_rows:
        new_df = pd.DataFrame(new_rows).set_index('date', drop=True)
        if output is not None:
            output = pd.concat([output, new_df])
        else:
            output = new_df

        output = output.sort_index()
        output.to_excel(output_path)

    return len(new_rows)


def update_xlsx_files(ticker: str, lktype: str):
    if lktype:
        spec_file = f'{ticker}/spec_{lktype}.xlsx'
        output_path = f'{ticker}/{ticker}-{lktype}.csv'
    else:
        spec_file = f'{ticker}/spec.xlsx'
        output_path = f'{ticker}/{ticker}.csv'


    if not os.path.exists(f'{ticker}/spec_xlsx.xlsx'):
        return
    
    spec = pd.read_excel(spec_file)
    spec['force_sign'] = spec['force_sign'].fillna(0).astype(int)

    files = glob.glob(f'{ticker}/20*.xlsx')
    files = sorted(files)

    if os.path.exists(output_path):
        output = pd.read_csv(output_path, parse_dates=[]).set_index('date', drop=True)
    else:
        output = None

    new_rows = []
    for file in files:
        date = os.path.basename(file).split('.')[0]
        if output is not None and date in output.index:
            continue

        row = parse_xlsx_file(spec, file)
        if output is not None and row['date'] in output.index:
            # Already exists
            pass
        else:
            print(f'Adding LK {row["date"]}')
            new_rows.append(row)

    if new_rows:
        new_df = pd.DataFrame(new_rows).set_index('date', drop=True)
        if output is not None:
            output = pd.concat([output, new_df])
        else:
            output = new_df

        output = output.sort_index()
        output.to_csv(output_path)

    return len(new_rows)


def get_links(html: str, pattern: str) -> List[str]:
    soup = BeautifulSoup(html, 'html.parser')
    links = soup.find_all('a')
    urls = []
    for link in links:
        href = link.get('href')
        if re.search(pattern, href):
            urls.append(href)

    return sorted(urls)


def check_web(ticker: str, update=False):
    with open(f'{ticker}/web.json') as f:
        spec = json.load(f)

    if "url" not in spec:
        return
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:125.0) Gecko/20100101 Firefox/125.0',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.7,id;q=0.3',
        'Accept-Encoding': 'gzip, deflate, br',
        'DNT': '1',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'none',
        'Sec-Fetch-User': '?1',
    }
    r = requests.get(spec['url'], headers=headers)
    web_links = get_links(r.text, spec['pattern'])
    if not web_links:
        raise RuntimeError(f'URL contains no PDF links: {spec["url"]}')
    
    path = f'{ticker}/laporan.html'

    if update:
        with open(path, 'wt') as f:
            f.write(r.text)
    else:
        if os.path.exists(path):
            with open(path) as f:
                html = f.read()
            saved_links = get_links(html, spec['pattern'])
        
            new_links = [l for l in web_links if l not in saved_links]
        else:
            new_links = web_links

        if new_links:
            print(f'{len(new_links)} new PDF file(s) detected:')
            for link in new_links:
                print('-', link)


if __name__ == '__main__':
    update = False
    argv = []
    lktype = ''

    i = 0
    while i < len(sys.argv):
        if sys.argv[i]=='-u':
            update = True
        elif sys.argv[i]=='-t':
            lktype = sys.argv[i+1]
            i += 1
        else:
            argv.append(sys.argv[i])
        i += 1

    check_web(argv[1], update=update)
    update_text_files(argv[1], lktype)
    update_xlsx_files(argv[1], lktype)
